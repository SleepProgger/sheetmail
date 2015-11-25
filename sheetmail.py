from openpyxl import load_workbook
from openpyxl.utils.exceptions import CellCoordinatesException, InvalidFileException

import smtplib
from email.mime.text import MIMEText 

import json
import logging
import argparse
import socket
from exceptions import ValueError
from time import sleep, time as now
from math import ceil
import os

log_debug, log_info, log_warn, log_error = logging.debug, logging.info, logging.warn, logging.error

def excel_data_iterator(workbook, sheet_index=0, selected_rows=tuple([0, 5]), row_offset=1):
    workbook.active = sheet_index
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows()):
        if row_number < row_offset: continue
        yield (row_number,) + tuple([row[i].value for i in selected_rows])


class Mail_sender():
    def __init__(self, server_config, quota_persist=None, nosend=False):
        self.host = server_config['host']
        self.port = server_config['port']
        self.username = server_config['username']
        self.password = server_config['password']
        self.sender = server_config['sender_addr']
        self.use_ssl = server_config['user_ssl']
        self.config = server_config
        self.nosend = nosend
        
        self.connected = False
        self.quota_persist = quota_persist
        self.next_send = 0
        self._update_quotas(mail_send=False)
        
    def _update_quotas(self, mail_send=True):
        curtime = now()
        if curtime >= self.config['timeframe_end']:
            self.config['timeframe_end'] = now() + self.config['timeframe']
            self.config['remaining_requests'] = self.config['allowed_requests']
            self.next_send = now() + self.config['timeframe']
            log_debug('New timeframe for %s@%s:%i. %i Requests till %s' % (self.username, self.host, self.port, self.config['remaining_requests'], self.config['timeframe_end']))
        # TODO: check if send was allowed ?
        if mail_send:
            self.config['remaining_requests'] -= 1            
        if self.config['update_config']:
            self.quota_persist()
        if self.config['use_fixed_delay']:
            if mail_send:
                self.next_send = curtime + ceil(self.config['timeframe'] / self.config['allowed_requests'])
            else:
                self.next_send = curtime
        else:    
            if self.config['remaining_requests'] > 0:
                mail_delay = ceil( (self.config['timeframe_end'] - curtime) / self.config['remaining_requests'] )
                log_debug('Can send 1 mail every %i second for %s@%s:%i. %i requests remaining in frame.' % (mail_delay, self.username, self.host, self.port, self.config['remaining_requests']))
                self.next_send = curtime + (mail_delay if mail_send else 0)
            else:
                self.next_send = self.config['timeframe_end']
        
            
    def connect_to_server(self):
        log_debug("Connecting to server %s@%s:%i." % (self.username, self.host, self.port))
        if self.use_ssl:
            self.server = smtplib.SMTP_SSL(self.host, self.port)
            self.server.ehlo() # optional, called by login()
        else:
            self.server = smtplib.SMTP(self.host, self.port)
            self.server.ehlo()
            self.server.starttls()
        self.server.login(self.username, self.password)  
        log_info('Logged in to %s:%i as %s' % (self.host, self.port, self.username))
        self.connected = True

    # TODO: unused atm. remove ?
    def can_send(self):
        return now() >= self.next_send

    def _send_mail(self, mail_to, mail_subject, mail_body):
        if not self.connected:
            self.connect_to_server()
        if isinstance(mail_to, (str, unicode)):
            mail_to = [mail_to]

        msg = MIMEText(mail_body, "plain", "utf-8")
        msg['Subject'] = mail_subject
        msg['From'] = self.sender
        msg['To'] = ",".join(mail_to)
         
        if not self.nosend:
            self.server.sendmail(self.sender, mail_to, msg.as_string())
        log_info('Send mail to %s over %s@%s:%i' % (mail_to, self.username, self.host, self.port))
        self._update_quotas(mail_send=True)
      
    def send_mail(self, mail_to, mail_subject, mail_body, retries=3, sleep_time=5):
        """
        Tries to send the mail trying to resolve some errors (up to `retries` times).
        If this function returns false, the connection is bugged and should be closed.
        May throw Exceptions:
        - smtplib.SMTPRecipientsRefused   Wrong recipient. Valid recipients might work.
        """
        for i in xrange(retries):
            try:
                self._send_mail(mail_to, mail_subject, mail_body)
                log_info('Successful send mail from %s to %s (Subject: %s) at %i try.' % (self.sender, mail_to, mail_subject, i+1))
                return True
            except smtplib.SMTPRecipientsRefused as e:
                # Avoid this being catched by the SMPTException catcher further down
                raise e  
            except smtplib.SMTPSenderRefused as e:
                log_error('The server refused the sender (%s) %s@%s:%i' % (self.sender, self.username, self.host, self.port))
                return False   
            except smtplib.SMTPAuthenticationError as e:
                log_error('SMTP authentication problem for connection %s@%s:%i' % (self.username, self.host, self.port))
                return False   
            except smtplib.SMTPHeloError as e:
                log_error('The server didn\'t reply properly to the HELO greeting.')
                return False
            except smtplib.SMTPDataError as e:
                log_warn('The server replied with an unexpected error code.')
                if i < retries:
                    log_info('Trying to reconnect to %s@%s:%i ...' % (self.username, self.host, self.port))
                    self.close()
                    sleep(sleep_time)
                    self.connected = False
            except socket.error as e:
                log_warn('Socket error: %s .' % e)
                if i < retries:
                    log_info('Trying to reconnect to %s@%s:%i ...' % (self.username, self.host, self.port))
                    self.close()
                    sleep(sleep_time)
                    self.connected = False
            except smtplib.SMTPConnectError as e:
                log_warn('Connection error for %s@%s:%i.' % (self.username, self.host, self.port))
                if i < retries:
                    log_info('Trying to reconnect to %s@%s:%i ...' % (self.username, self.host, self.port))
                    self.close()
                    sleep(sleep_time)
                    self.connected = False
            except smtplib.SMTPException as e:
                log_error('Smtplib error: %s .' % e)
                return False
        log_error('To many retries sending to %s:%i via %s. Giving up.' % (self.host, self.port, self.username))
        return False
    
    def close(self):
        try: # DealWithIt.mov
            self.server.quit()
        except: pass
        self.connected = False



class Excel_Mail_Sender():
    def __init__(self, config):
        self.config = config;
        self.server_config = None
        self.server_connections = None
        self.wb = None
        self.sheet = None
        
    def init(self):
        try:
            log_debug("Open '%s'" % self.config['excel_file'])
            self.wb = load_workbook(self.config['excel_file'], data_only=True, keep_vba=True)
        except CellCoordinatesException, InvalidFileException:
            log_error('Failed to load spreadsheet file "%s"' % self.config['excel_file'])
            return False
        
        if self.config['cleancomments']:
            log_info("Clean comments from spreadsheet to avoid comment bug.")
            for i in xrange(len(self.wb.get_sheet_names())):
                self.wb.active = i
                for row in self.wb.active.rows:
                    for cell in row: cell.comment = None
            
        self.wb.active = self.config['sheetindex']
        self.sheet = self.wb.active
        log_info('Loaded excel file "%s"' % self.config['excel_file'])
        try:
            self.server_config = json.load(self.config['config'])
            if not 'mail_user' in self.server_config:
                raise ValueError()
            log_debug('Parsed config file "%s".' % self.config['config'].name)
        except ValueError:
            log_error('Failed to parse config "%s".' % self.config['config'].name)
            return False
        
        connections = list()
        for connection in self.server_config['mail_user']:
            sender = Mail_sender(connection, self._persist_config, nosend=self.config['nosend'])
            # TODO: should we try to connect with every user once here ?
            log_info('Loaded data for connection %s@%s:%i' % (sender.username, sender.host, sender.port))
            connections.append(sender)
        self.server_connections = connections
        return True
        
    def _persist_config(self):
        self.config['config'].seek(0)
        json.dump(self.server_config, self.config['config'], indent=True, sort_keys=True)
        self.config['config'].truncate()
        self.config['config'].flush() # Do we need this here ?
    
    def _find_free_server(self):
        a = sorted(((x.next_send, x) for x in self.server_connections))
        return a[0]
    
    def close(self):
        self.wb.save(self.config['excel_file'])
    
    def run(self):
        row_select = (self.config['colmail'], self.config['colsubject'], self.config['colbody'], self.config['colsend'])
        for row in excel_data_iterator(self.wb, self.config['sheetindex'], row_select, self.config['rowoffset']):
            row = list(row)
            if 'staticsubject' in self.config:
                row[2] = self.config['staticsubject']
            if row[4] and row[4] >= 1:
                log_debug('Skip row %i, because it was already send.' % row[0])
                continue
            
            next_send, connection = self._find_free_server()
            if now() < next_send:
                _next = next_send - now()
                log_debug('Next mail send in %i seconds' % _next)
                sleep(_next)
            
            failed = False
            if not row[1] or not row[2] or not row[3]:
                log_error('Invalid data in row %i. Skipping row.' % (row[0] + 1))
                failed = True
            else:
                try:
                    status = connection.send_mail(row[1], row[2], row[3], retries=3, sleep_time=5) 
                except smtplib.SMTPRecipientsRefused as e:
                    log_error('Recipients refused (%s) for connection %s@%s:%i' % (row[1], connection.username, connection.host, connection.port))
                    failed = True
            if failed:
                self.sheet.cell(row=row[0] + 1, column=self.config['colsend'] + 1).value = 2
                self.wb.save(self.config['excel_file'])
                continue
                
            if status:
                self.sheet.cell(row=row[0] + 1, column=self.config['colsend'] + 1).value = 1
                self.wb.save(self.config['excel_file'])
            else:
                # We do NOT set the error or done flag when there is an critical error
                # as it is probably some network or configuration error- 
                log_error('Critical error. Going down.')
                return False


def test_mail(config):
    server_config = json.load(config['config'])
    for connection in server_config['mail_user']:
        log_info('Try to connect to %s@%s:%i ...' % (connection['username'], connection['host'], connection['port']))
        sender = Mail_sender(connection, None)
        try:
            sender.connect_to_server()
            log_info('Connection valid.')
            sender.close()
        except smtplib.SMTPHeloError:
            log_error('The server didn\'t reply properly to the HELO greeting.')
        except smtplib.SMTPAuthenticationError:
            log_error('The server didn\'t accept the username/password combination.')
        except smtplib.SMTPException:
            log_error('No suitable authentication method was found.')
        except socket.error as e:
            log_error('Network error: %s' % e)
          
def test_spreadsheet_file(config, cleancomments):
    from tempfile import mkstemp
    error = False
    log_info('Trying to load spreadsheet file "%s" ...' % config['excel_file'])
    try:
        book = load_workbook(config['excel_file'], use_iterators=False, keep_vba=True)
    except InvalidFileException:
        log_error('Invalid file. Can\'t open')
        return True
        
    if cleancomments:
        log_info('Cleaning comments for temporary file')
        for i in xrange(len(book.get_sheet_names())):
            book.active = i
            for row in book.active.rows:
                for cell in row: cell.comment = None
    suffix = '.' + config['excel_file'].split('.')[-1]
    book.active = config['sheetindex']
    # Sometimes max_row isn't correct, so we shouldn't rely on it. TODO: it would be great if we could
    log_info('Loading succesfully. %i rows found.' % book.active.max_row)
    
    # https://bitbucket.org/openpyxl/openpyxl/issues/536/cant-save-and-reopen-xlsx-file-with
    log_debug('Checking for comment bug ...')
    tmp_file, tmp_name = mkstemp(suffix=suffix)
    os.close(tmp_file)
    book.save(tmp_name)
    fd = open(tmp_name, 'rb')
    try:
        book = load_workbook(filename=fd, use_iterators=False, keep_vba=True)
        book.active = config['sheetindex']
        log_info('Comment bug not detected. Rows %i' % book.active.max_row)
    except TypeError as e:
        log_error('Failed reloading file. Try again with added --cleancomments parameter.')
        error = True
    fd.close()
    os.remove(tmp_name)
    return error 


if __name__ == '__main__':
    # TODO: CSV ? The default lib has unicode problems
    import sys
    parser = argparse.ArgumentParser(description='Sends emails with data supplied by excel files.')
    parser.add_argument('--config', '-c', required=True, type=argparse.FileType('r+b'), default='./config.json', help='Choose the configuration file.')
    parser.add_argument('--loglvl', '-l', help='Set the log level.', default='INFO', choices=('DEBUG', 'INFO', 'WARN', 'ERROR'))
    parser.add_argument('--logfile', '-f', type=argparse.FileType('w'), help='Also write log to file.')    
    parser.add_argument('--rowoffset', '-r', type=int, default=1, help='The row to start with.')
    parser.add_argument('--colmail', '-m', type=int, default=0, help='The column containing the email address.')
    parser.add_argument('--colsubject', '-s', type=int, default=1, help='The column containing the email subjects.')
    parser.add_argument('--colbody', '-b', type=int, default=2, help='The column containing the email message.')
    parser.add_argument('--colsend', '-o', type=int, default=3, help='The column used to mark if the mail was send.')
    parser.add_argument('--staticsubject', '-x', help='Can be used to use a static subject.')
    parser.add_argument('--sheetindex', '-i', type=int, default=0, help='The sheet to use.')
    parser.add_argument('--cleancomments', action='store_true', help='Remove comments from file. Openpyxl has/hadd a bug leading to corrupt files otherwise.')
    parser.add_argument('excel_file', type=argparse.FileType('r+b'), help='The excel file to get data from.')
    parser.add_argument('--test', action='store_true', help='Only test all mail accounts and the spreadsheet file. Then exit.')       
    parser.add_argument('--nosend', action='store_true', help='Do NOT send the mails. Used for testing.')       
    parser.add_argument('--notest', action='store_true', help='Do NOT test for comment bug on startup.')       
    params = vars(parser.parse_args(sys.argv[1:]))
    
    # Get filename of the excel file (used to check perms)
    file_name = params['excel_file'].name
    params['excel_file'].close()
    params['excel_file'] = file_name
    
    # Set up logging
    if params['loglvl'] == 'DEBUG':
        log_lvl = logging.DEBUG 
    elif params['loglvl'] == 'INFO':
        log_lvl = logging.INFO 
    elif params['loglvl'] == 'WARN':
        log_lvl = logging.WARN 
    elif params['loglvl'] == 'ERROR':
        log_lvl = logging.ERROR
    if 'logfile' in params:
        logging.basicConfig(filename=params['logfile'], format='%(asctime)s %(levelname)s: %(message)s', datefmt='%m/%d %I:%M %p', level=log_lvl)
    else:
        logging.basicConfig(level=log_lvl, format='%(asctime)s %(levelname)s: %(message)s', datefmt='%m/%d %I:%M %p')
    
    if params['test']:
        test_mail(params)
        test_spreadsheet_file(params, params['cleancomments'])
        exit()
        
    if not params['notest']:
        log_info('Testing file integrity and comment bug...')
        if test_spreadsheet_file(params, params['cleancomments']):
            exit()
    
    sender = Excel_Mail_Sender(params)
    if sender.init():
        try:
            sender.run()
        except KeyboardInterrupt:
            sender.close()
